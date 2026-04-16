from __future__ import annotations

import csv
from collections import OrderedDict
from functools import cached_property
from io import BytesIO
from typing import TYPE_CHECKING

from django.core.exceptions import ValidationError
from django.db.models import Count
from django.http import FileResponse, StreamingHttpResponse
from django.utils.translation import gettext_lazy as _
from wagtail.admin.panels import (
    FieldPanel,
    MultiFieldPanel,
    ObjectList,
)
from wagtail.admin.ui.components import Component
from wagtail.admin.ui.menus import MenuItem
from wagtail.admin.ui.tables import Column
from wagtail.admin.views.mixins import Echo
from wagtail.admin.widgets.button import Button
from wagtail.images.widgets import AdminImageChooser
from wagtail.snippets.models import register_snippet

from dal import autocomplete
from openpyxl import Workbook

from kausal_common.users import user_or_bust

from admin_site.forms import WatchAdminModelForm
from admin_site.permissions import PlanRelatedPermissionPolicy
from admin_site.viewsets import WatchCreateView, WatchEditView, WatchIndexView, WatchViewSet
from admin_site.wagtail import (
    AplansTabbedInterface,
    get_translation_tabs,
)

from .models import Pledge
from .models.pledge import PledgeCommitment

if TYPE_CHECKING:
    from collections.abc import Generator

    from django.contrib.auth.base_user import AbstractBaseUser
    from django.contrib.auth.models import AnonymousUser
    from django.http import HttpRequest
    from django.utils.safestring import SafeString


class PledgePermissionPolicy(PlanRelatedPermissionPolicy):
    """Permission policy for Pledges - only accessible if community engagement is enabled."""

    def user_has_permission(self, user: AbstractBaseUser | AnonymousUser, action: str) -> bool:
        if not super().user_has_permission(user, action):
            return False

        # Check if community engagement is enabled
        from users.models import User

        if not isinstance(user, User):
            return False

        plan = user.get_active_admin_plan(required=False)
        if plan is None:
            return False

        return plan.features.enable_community_engagement


class PledgeAdminForm(WatchAdminModelForm[Pledge]):
    """Form for editing Pledges with attribute support."""

    class Meta:
        model = Pledge
        # Exclude auto-generated, relationship, and i18n storage fields
        # The i18n fields are handled by translation tabs via get_translation_tabs()
        exclude = [
            'uuid',
            'plan',
            'order',
            'name_i18n',
            'description_i18n',
            'impact_statement_i18n',
            'local_equivalency_i18n',
        ]
        # Explicitly define widgets to ensure they're used even with dynamic form class creation
        widgets = {
            'image': AdminImageChooser,
            'actions': autocomplete.ModelSelect2Multiple(url='action-autocomplete'),
        }

    def clean_slug(self):
        # Since the plan field is excluded from the form, `validate_unique()` won't check
        # the unique_together = [('plan', 'slug')] constraint. We validate it manually here.
        slug = self.cleaned_data['slug']
        plan = self.instance.plan
        if Pledge.objects.filter(plan=plan, slug=slug).exclude(pk=self.instance.pk).exists():
            raise ValidationError(_('There is already a pledge with this slug.'))
        return slug

    def save(self, commit=True):
        obj = super().save(commit)
        # Get user from the dynamically set _user attribute
        user = getattr(self, '_user', None)
        if user is None:
            # Fallback: no attribute saving if user is not set
            return obj
        # Save attribute values from the form
        if commit:
            attribute_types = obj.get_editable_attribute_types(user)
        else:
            attribute_types = obj.get_visible_attribute_types(user)
        for attribute_type in attribute_types:
            attribute_type.on_form_save(obj, self.cleaned_data, commit=commit)
        return obj


class PledgeViewMixin:
    """Mixin for Pledge create/edit views with dynamic attribute panels."""

    model: type[Pledge]
    request: HttpRequest

    def get_form_class(self):
        """Build form class with dynamic attribute fields."""
        from kausal_common.users import user_or_bust

        request = self.request
        user = user_or_bust(request.user)
        plan = user.get_active_admin_plan()

        # Get instance - if editing use self.object, if creating make temporary instance
        instance: Pledge | None = getattr(self, 'object', None)
        if not instance or not instance.pk:
            # For create view, make a temporary instance with plan set
            instance = self.model(plan=plan)  # type: ignore[call-arg]

        # Get attribute types and fields
        attribute_types = instance.get_visible_attribute_types(user)
        attribute_fields = {
            field.name: field.django_field
            for attribute_type in attribute_types
            for field in attribute_type.get_form_fields(user, plan, instance)
        }

        # Create a dynamic form class with attribute fields and user context.
        # Include _user as a class attribute so the form can access it.
        form_attrs = {**attribute_fields, '_user': user}
        form_class = type(
            'DynamicPledgeAdminForm',
            (PledgeAdminForm,),
            form_attrs,
        )
        return form_class

    def get_panel(self):
        """Return edit handler with dynamically built panels including attributes."""
        from kausal_common.users import user_or_bust

        request = self.request
        user = user_or_bust(request.user)
        plan = user.get_active_admin_plan()

        # Get the instance - if editing, use self.object; if creating, make a temporary instance with plan
        instance: Pledge
        if hasattr(self, 'object') and self.object and self.object.pk:
            instance = self.object
        else:
            # For create view, make a temporary instance with plan set so attribute types can be queried
            instance = self.model(plan=plan)  # type: ignore[call-arg]

        # Get attribute panels
        main_attribute_panels, i18n_attribute_panels = instance.get_attribute_panels(user)

        # Build panels list using ViewSet's panels directly (don't copy to preserve widget config)
        # Insert attribute panels at the position specified by attribute_panel_position
        pos = PledgeViewSet.attribute_panel_position
        panels: list = list(PledgeViewSet.panels[:pos])

        # Add attribute panels
        if main_attribute_panels:
            panels.extend(list(main_attribute_panels))

        # Add remaining panels
        panels.extend(PledgeViewSet.panels[pos:])

        # Get translation tabs
        i18n_tabs = get_translation_tabs(instance, request, extra_panels=i18n_attribute_panels)

        # If there are translation tabs, use TabbedInterface; otherwise just ObjectList
        if i18n_tabs:
            tabs = [
                ObjectList(panels, heading=_('Basic information')),
            ] + i18n_tabs
            return AplansTabbedInterface(tabs).bind_to_model(self.model)
        return ObjectList(panels).bind_to_model(self.model)


class PledgeCreateView(PledgeViewMixin, WatchCreateView[Pledge]):
    """Custom create view for Pledge with dynamic attribute panels."""


class PledgeEditView(PledgeViewMixin, WatchEditView[Pledge]):
    """Custom edit view for Pledge with dynamic attribute panels."""


class _DropdownLabel(MenuItem, Component):
    """Non-interactive group label rendered inside a dropdown."""

    def __init__(self, label: str, priority: int = 0) -> None:
        super().__init__(label, '', icon_name = '', priority=priority)

    def render_html(self, parent_context=None) -> SafeString:
        from django.utils.html import format_html
        return format_html(
            '<div style="padding: 0.5rem 1.5rem 0.25rem; font-size: 0.75rem; font-weight: 600;'
            ' text-transform: uppercase; color: var(--w-color-text-label-menus-default); opacity: 0.7;">{}</div>',
            self.label,
        )


class _DropdownDivider(MenuItem, Component):
    """Horizontal rule rendered inside a dropdown."""

    def __init__(self, priority: int = 0) -> None:
        super().__init__('', '', icon_name = '', priority=priority)

    def render_html(self, parent_context=None) -> SafeString:
        from django.utils.safestring import mark_safe
        return mark_safe(
            '<hr style="border: none; border-top: 1px solid var(--w-color-border-furniture); margin: 0.25rem 0;">'
        )


class PledgeIndexView(WatchIndexView[Pledge]):
    """Custom index view for Pledge with spreadsheet export support."""

    export_headings = {
        'commitment_count': _('Number of commitments'),
    }
    show_export_buttons = False

    @property
    def list_export(self) -> list[str]:
        return ['id', 'name', 'slug', 'commitment_count']

    @list_export.setter
    def list_export(self, value: list[str]) -> None:
        pass

    @property
    def export_filename(self) -> str:
        plan = user_or_bust(self.request.user).get_active_admin_plan()
        return f'{_("Pledges")} - {plan}'

    @export_filename.setter
    def export_filename(self, value: str) -> None:
        pass

    def get_header_more_buttons(self):
        buttons = list(super().get_header_more_buttons())

        def _url(export_type: str, fmt: str) -> str:
            params = self.request.GET.copy()
            params['export'] = fmt
            params['export_type'] = export_type
            return self.request.path + '?' + params.urlencode()

        buttons += [
            _DropdownLabel(str(_('Export')), priority=89),
            Button(_('Export as Excel'),              url=self.get_export_url('xlsx'), icon_name='download', priority=90),
            Button(_('Export pledges as CSV'),        url=_url('pledges',     'csv'),  icon_name='download', priority=91),
            Button(_('Export commitments as CSV'),    url=_url('commitments', 'csv'),  icon_name='download', priority=92),
            _DropdownDivider(priority=93),
        ]
        return sorted(buttons)

    @cached_property
    def _user_data_keys(self) -> list[str]:
        """Discover all unique keys from user_data across commitments for the current queryset."""
        qs = self.get_queryset()
        # Wagtail search returns PostgresSearchResults which doesn't support values_list
        if hasattr(qs, 'values_list'):
            pledge_ids = list(qs.values_list('pk', flat=True))
        else:
            pledge_ids = [item.pk for item in qs]
        user_data_values = (
            PledgeCommitment.objects
            .filter(pledge_id__in=pledge_ids)
            .exclude(pledge_user__user_data={})
            .values_list('pledge_user__user_data', flat=True)
        )
        return sorted({key for user_data in user_data_values if user_data for key in user_data})

    def stream_csv(self, queryset):
        """Override to use QUOTE_ALL so fields with spaces aren't split into separate columns."""
        writer = csv.DictWriter(Echo(), fieldnames=self.list_export, quoting=csv.QUOTE_ALL)
        yield writer.writerow({field: self.get_heading(queryset, field) for field in self.list_export})
        for item in queryset:
            yield self.write_csv_row(writer, self.to_row_dict(item))

    def to_row_dict(self, item: Pledge) -> OrderedDict[str, str]:
        return OrderedDict([
            ('id', str(item.pk)),
            ('name', str(item.name)),
            ('slug', item.slug),
            ('commitment_count', str(getattr(item, 'commitment_count', item.commitments.count()))),
        ])

    def render_to_response(self, context, **response_kwargs):
        if (self.is_export
                and self.request.GET.get('export_type') == 'commitments'
                and self.request.GET.get('export') == 'csv'):
            return self.write_commitments_csv_response(context['object_list'])
        return super().render_to_response(context, **response_kwargs)

    @property
    def _commitment_export_fields(self) -> list[str]:
        return ['pledge_id', 'pledge_name', 'commitment_created_at', 'user_id'] + [
            f'user_data:{key}' for key in self._user_data_keys
        ]

    @property
    def _commitment_export_headings(self) -> dict[str, str]:
        headings: dict[str, str] = {
            'pledge_id': str(_('Pledge ID')),
            'pledge_name': str(_('Pledge name')),
            'commitment_created_at': str(_('Commitment date')),
            'user_id': str(_('User ID')),
        }
        for key in self._user_data_keys:
            headings[f'user_data:{key}'] = key
        return headings

    def _commitment_export_rows(self, queryset) -> Generator[dict[str, str]]:
        """Yield one dict per commitment across all pledges in the queryset."""
        for pledge in queryset:
            for commitment in pledge.commitments.select_related('pledge_user').order_by('created_at'):
                user_data = commitment.pledge_user.user_data or {}
                row: dict[str, str] = {
                    'pledge_id': str(pledge.pk),
                    'pledge_name': str(pledge.name),
                    'commitment_created_at': commitment.created_at.isoformat(),
                    'user_id': str(commitment.pledge_user.uuid),
                }
                for key in self._user_data_keys:
                    row[f'user_data:{key}'] = str(user_data[key]) if key in user_data else ''
                yield row

    def _commitment_export_filename(self) -> str:
        plan = user_or_bust(self.request.user).get_active_admin_plan()
        return f'{_("Pledge Commitments")} - {plan}'

    def write_commitments_csv_response(self, queryset) -> StreamingHttpResponse:
        fields = self._commitment_export_fields
        headings = self._commitment_export_headings

        def _stream() -> Generator[bytes]:
            writer = csv.DictWriter(Echo(), fieldnames=fields, quoting=csv.QUOTE_ALL)
            yield writer.writerow(headings)
            for row in self._commitment_export_rows(queryset):
                yield writer.writerow(row)

        response = StreamingHttpResponse(_stream(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{self._commitment_export_filename()}.csv"'
        return response

    def write_xlsx_response(self, queryset) -> FileResponse:
        output = BytesIO()
        workbook = Workbook(write_only=True, iso_dates=True)

        # Pledges sheet
        pledges_sheet = workbook.create_sheet(title=str(_('Pledges')))
        pledges_sheet.append([self.get_heading(queryset, f) for f in self.list_export])
        for item in queryset:
            row = self.to_row_dict(item)
            pledges_sheet.append([row[f] for f in self.list_export])

        # Commitments sheet
        commitment_fields = self._commitment_export_fields
        commitment_headings = self._commitment_export_headings
        commitments_sheet = workbook.create_sheet(title=str(_('Commitments')))
        commitments_sheet.append([commitment_headings[f] for f in commitment_fields])
        for row in self._commitment_export_rows(queryset):
            commitments_sheet.append([row[f] for f in commitment_fields])

        workbook.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=f'{self.export_filename}.xlsx',
        )


class PledgeViewSet(WatchViewSet[Pledge]):
    """Admin interface for Pledges."""

    model = Pledge
    menu_label = _('Pledges')
    icon = 'kausal-pledge'
    menu_icon = 'kausal-pledge'
    menu_order = 41  # After Indicators (40)
    add_to_admin_menu = True
    list_display = [
        'name',
        Column('commitment_count', label=_('Commitments'), sort_key='commitment_count'),
        Column('updated_at', label=_('Updated at'), sort_key='updated_at'),
    ]
    search_fields = ['name']
    ordering = ['plan', 'order']
    index_view_class = PledgeIndexView  # type: ignore[assignment]
    add_view_class = PledgeCreateView  # type: ignore[assignment]
    edit_view_class = PledgeEditView  # type: ignore[assignment]

    # Define base panels to prevent Wagtail from auto-generating form fields
    # These are overridden by PledgeEditView.get_panel() which dynamically constructs
    # panels including attributes
    panels = [
        FieldPanel('name'),
        FieldPanel('slug'),
        FieldPanel('description'),
        FieldPanel('image'),
        # Attribute panels will come here (set in PledgeViewSet.attribute_panel_position)
        FieldPanel('body'),
        MultiFieldPanel(
            [
                FieldPanel('resident_count'),
                FieldPanel('impact_statement'),
                FieldPanel('local_equivalency'),
            ],
            heading=_('Community Impact Visualization'),
        ),
        FieldPanel('actions', widget=autocomplete.ModelSelect2Multiple(url='action-autocomplete')),
    ]

    # Position in `panels` where attribute panels are inserted (after name, slug, description, image)
    attribute_panel_position = 4

    @property
    def permission_policy(self):
        return PledgePermissionPolicy(self.model)

    def get_menu_item(self, order=None):
        """Customize menu item to check feature flag for visibility."""
        menu_item = super().get_menu_item(order)

        def is_shown(request) -> bool:
            from users.models import User

            if not isinstance(request.user, User):
                return False

            plan = request.user.get_active_admin_plan(required=False)
            if plan is None:
                return False

            return plan.features.enable_community_engagement

        menu_item.is_shown = is_shown  # type: ignore[method-assign]
        return menu_item

    def get_queryset(self, request: HttpRequest):
        qs = super().get_queryset(request)
        if qs is None:
            qs = self.model._default_manager.all()
        user = user_or_bust(request.user)
        plan = user.get_active_admin_plan()
        return qs.filter(plan=plan).annotate(commitment_count=Count('commitments'))


register_snippet(PledgeViewSet)
